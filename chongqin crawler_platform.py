import json
import time
import os
import re
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

# ===================== 配置参数 =====================
TARGET_URL = "https://www.cqggzy.com/jyxx/transaction_detail.html"
BASE_URL = "https://www.cqggzy.com"
# 保持原有字段完全兼容
FIELDS = [
    "标题", "链接", "发布日期", "省份", "来源平台",
    "业务类型", "信息类型", "行业"
]
# 页面等待与重试配置
PAGE_WAIT = 2
MAX_RETRY = 3
ELEMENT_WAIT_TIMEOUT = 15  # 元素加载超时时间，适配网络波动
# 业务类型编码映射（从页面HTML提取，自动填充空字段）
BUSINESS_TYPE_MAP = {
    "014001": "工程招投标",
    "014005": "政府采购",
    "014012": "国企物资采购",
    "014008": "其他采购",
    "014010": "中介超市",
    "014002": "产权交易",
    "014014": "农村产权",
    "014004": "土地及矿业权",
    "014011": "药械交易",
    "014006": "碳排放权交易",
    "014009": "排污权交易",
    "014003": "机电设备",
    "014013": "诉讼资产交易"
}
# 信息类型编码映射
INFO_TYPE_MAP = {
    "001": "招标公告",
    "002": "答疑补遗/变更",
    "003": "中标候选人公示",
    "004": "中标结果公示",
    "005": "采购公告",
    "019": "招标计划",
    "014": "邀标信息",
    "020": "合同签订公示",
    "021": "终止公告",
    "008": "单一来源公示"
}

# ===================== 无头浏览器初始化（反爬增强）=====================
def init_driver():
    """初始化无头Edge浏览器，适配GitHub Actions Ubuntu环境，绕过反爬检测"""
    edge_options = Options()
    
    # 无头模式核心配置
    edge_options.add_argument("--headless=new")  # Edge新版无头模式，更难被检测
    edge_options.add_argument("--no-sandbox")  # Ubuntu环境必须参数
    edge_options.add_argument("--disable-dev-shm-usage")  # 解决共享内存不足问题
    edge_options.add_argument("--disable-gpu")  # 无头环境禁用GPU
    edge_options.add_argument("--window-size=1920,1080")  # 固定窗口尺寸，防止分辨率检测
    
    # 反爬核心配置：隐藏自动化特征
    edge_options.add_argument("--disable-blink-features=AutomationControlled")
    edge_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    edge_options.add_experimental_option("useAutomationExtension", False)
    
    # 真实浏览器指纹配置
    edge_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0")
    edge_options.add_argument("--accept-language=zh-CN,zh;q=0.9,en;q=0.8")
    edge_options.add_argument("--disable-extensions")
    edge_options.add_argument("--disable-plugins-discovery")
    
    # 启动浏览器
    driver = webdriver.Edge(options=edge_options)
    driver.maximize_window()
    
    # 执行JS脚本，彻底隐藏自动化特征
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            // 隐藏webdriver属性
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            // 伪造插件信息
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            // 伪造语言信息
            Object.defineProperty(navigator, 'languages', {
                get: () => ['zh-CN', 'zh', 'en']
            });
            // 伪造chrome对象
            window.chrome = { runtime: {} };
            // 伪造权限查询
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({
                    query: () => Promise.resolve({ state: 'granted' })
                })
            });
            // 覆盖cdc属性
            Object.defineProperty(window, 'cdc_adoQpoasnfa76pfcZLmcfl_', {
                get: () => undefined
            });
        """
    })
    
    print("✅ 无头Edge浏览器已启动，反爬配置已加载")
    return driver

# ===================== 数据解析函数 =====================
def parse_notice_item(notice_soup):
    """解析单个公告条目，保持原有字段兼容，自动补全业务/信息类型"""
    item = {field: "" for field in FIELDS}
    item["省份"] = "重庆市"
    
    try:
        # 1. 标题、链接、来源平台提取
        a_tag = notice_soup.find("a", class_="l")
        if a_tag:
            # 标题提取
            item["标题"] = a_tag.get("title", "").strip() or a_tag.get_text(strip=True)
            # 链接补全
            href = a_tag.get("href", "").strip()
            if href.startswith("/"):
                item["链接"] = BASE_URL + href
            else:
                item["链接"] = href
            
            # 来源平台提取
            region_tag = a_tag.find("span", class_="region")
            if region_tag:
                item["来源平台"] = region_tag.get_text(strip=True).replace("【", "").replace("】", "")
            
            # 从链接路径提取业务类型、信息类型
            href_path = href.split("/")
            if len(href_path) >= 3:
                # 匹配业务类型编码
                business_code = href_path[2]
                for code, name in BUSINESS_TYPE_MAP.items():
                    if code in business_code:
                        item["业务类型"] = name
                        break
                # 匹配信息类型编码
                info_code = href_path[3] if len(href_path) >=4 else ""
                for code, name in INFO_TYPE_MAP.items():
                    if code in info_code:
                        item["信息类型"] = name
                        break
        
        # 2. 发布日期提取
        date_tag = notice_soup.find("span", class_="info-date")
        if date_tag:
            item["发布日期"] = date_tag.get_text(strip=True)
        
        # 行业字段保留兼容，可按需扩展
        item["行业"] = ""
                
    except Exception as e:
        print(f"⚠️ 解析单条数据失败: {e}")
    
    return item

def parse_current_page(driver):
    """解析当前页面所有公告数据，带重试机制"""
    soup = BeautifulSoup(driver.page_source, "html.parser")
    # 适配重庆市平台的列表容器
    notice_list = soup.select("ul#showList > li.info-item")
    page_data = []
    
    for notice in notice_list:
        item = parse_notice_item(notice)
        page_data.append(item)
    
    print(f"📄 当前页解析出 {len(page_data)} 条数据")
    return page_data

def has_next_page(driver):
    """判断是否存在下一页，适配重庆市平台分页结构"""
    try:
        next_btn_list = driver.find_elements(By.XPATH, "//div[@id='divInfoReportPage']//a[contains(text(),'下页')]")
        if next_btn_list:
            next_btn = next_btn_list[0]
            # 检查按钮是否可点击
            if next_btn.is_enabled() and "disabled" not in (next_btn.get_attribute("class") or ""):
                return next_btn
        return None
    except Exception as e:
        print(f"⚠️ 检测下一页失败: {e}")
        return None

# ===================== 数据导出函数 =====================
def export_to_excel(data, filename=None):
    """导出数据到Excel，统一保存到data/目录，适配GitHub Actions"""
    if not data:
        print("❌ 无数据可导出Excel")
        return
    
    # 生成默认文件名
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"重庆市公共资源交易数据_{timestamp}.xlsx"
    
    # 确保data目录存在
    os.makedirs("data", exist_ok=True)
    filepath = os.path.join("data", filename)
    
    # 创建工作簿与工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "爬取数据"
    
    # 设置表头样式
    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 写入表头
    for col, field in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    
    # 写入数据行
    for row, item in enumerate(data, 2):
        for col, field in enumerate(FIELDS, 1):
            ws.cell(row=row, column=col, value=item[field])
    
    # 自动调整列宽
    for col in range(1, len(FIELDS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # 保存文件
    wb.save(filepath)
    print(f"✅ Excel数据已导出到: {filepath}")
    return filepath

def export_to_json(data, filename=None):
    """导出数据到JSON文件，统一保存到data/目录，适配GitHub Actions"""
    if not data:
        print("❌ 无数据可导出JSON")
        return
    
    # 生成默认文件名
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"重庆市公共资源交易数据_{timestamp}.json"
    
    # 确保data目录存在
    os.makedirs("data", exist_ok=True)
    filepath = os.path.join("data", filename)
    
    # 写入JSON文件，中文不转义，格式化输出
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    
    print(f"✅ JSON数据已导出到: {filepath}")
    return filepath

# ===================== 主程序（全自动运行+新增停止条件）=====================
def main():
    print("="*60)
    print("📢 重庆市公共资源交易平台全自动爬虫启动（新增日期停止条件）")
    print("="*60)
    
    # 计算北京时间（适配GitHub Actions UTC时区）
    beijing_now = datetime.utcnow() + timedelta(hours=8)
    # 目标爬取日期：前一天
    target_date = (beijing_now - timedelta(days=1)).strftime("%Y-%m-%d")
    # 停止触发日期：前天，页面出现此日期数据立即停止爬取
    stop_date = (beijing_now - timedelta(days=2)).strftime("%Y-%m-%d")
    print(f"📅 目标爬取日期（前一天）: {target_date}")
    print(f"🛑 停止触发日期（前天）: {stop_date}")
    
    all_data = []
    current_page_num = 1
    consecutive_failures = 0
    driver = None
    crawl_stop_flag = False  # 爬取停止标记
    
    try:
        # 1. 初始化浏览器
        driver = init_driver()
        
        # 2. 导航到目标页面
        driver.get(TARGET_URL)
        print(f"🌐 已导航到目标网址: {TARGET_URL}")
        
        # 3. 等待页面核心元素加载完成
        WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, "showList"))
        )
        print("✅ 页面核心元素加载完成")
        
        # 4. 自动点击「近三天」按钮，缩小数据范围
        try:
            three_days_btn = WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@id='day']//a[@value='2']"))
            )
            # 点击近三天按钮
            three_days_btn.click()
            print("✅ 已点击「近三天」筛选按钮")
            
            # 等待页面列表刷新完成
            time.sleep(PAGE_WAIT)
            WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "ul#showList > li.info-item"))
            )
            print("✅ 数据列表已刷新")
            
        except Exception as e:
            print(f"⚠️ 点击筛选按钮失败，将使用全量数据过滤: {e}")
        
        # 5. 循环爬取页面，新增停止条件判断
        while True:
            if crawl_stop_flag:
                break
            
            print(f"\n🚀 正在爬取第 {current_page_num} 页...")
            
            # 解析当前页，带重试机制
            page_raw_data = []
            for retry in range(MAX_RETRY):
                try:
                    page_raw_data = parse_current_page(driver)
                    if page_raw_data:
                        consecutive_failures = 0
                        break
                    else:
                        print(f"⚠️ 第 {current_page_num} 页数据为空，重试 ({retry+1}/{MAX_RETRY})...")
                        time.sleep(PAGE_WAIT)
                except Exception as e:
                    print(f"⚠️ 解析第 {current_page_num} 页失败，重试 ({retry+1}/{MAX_RETRY}): {e}")
                    time.sleep(PAGE_WAIT)
            
            # 处理空数据情况
            if not page_raw_data:
                consecutive_failures += 1
                print(f"❌ 第 {current_page_num} 页多次尝试后仍无数据")
                if consecutive_failures >= MAX_RETRY:
                    print("🔚 连续多次无数据，终止爬取")
                    break
                continue
            
            # ========== 核心新增：停止条件判断 ==========
            # 检查当前页是否包含停止日期（前天）的数据
            has_stop_date = any(item["发布日期"] == stop_date for item in page_raw_data)
            if has_stop_date:
                print(f"🛑 检测到前天({stop_date})数据，触发停止条件，保存当前页有效数据后终止爬取")
                # 过滤当前页符合目标日期的数据，加入总数据
                target_data = [item for item in page_raw_data if item["发布日期"] == target_date]
                all_data.extend(target_data)
                print(f"🔍 本页共 {len(page_raw_data)} 条，目标日期({target_date})有效数据 {len(target_data)} 条")
                crawl_stop_flag = True
                break
            
            # 无停止日期，正常过滤目标数据
            target_data = [item for item in page_raw_data if item["发布日期"] == target_date]
            all_data.extend(target_data)
            print(f"🔍 本页共 {len(page_raw_data)} 条，目标日期({target_date})有效数据 {len(target_data)} 条")
            
            # 检查是否有下一页
            next_btn = has_next_page(driver)
            if next_btn:
                try:
                    # 点击下一页
                    next_btn.click()
                    time.sleep(PAGE_WAIT)
                    # 等待页面列表加载完成
                    WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                        lambda d: len(d.find_elements(By.CSS_SELECTOR, "ul#showList > li.info-item")) > 0
                    )
                    current_page_num += 1
                except Exception as e:
                    print(f"⚠️ 点击下一页失败: {e}")
                    break
            else:
                print("🔚 已爬取到最后一页，无更多数据")
                break
        
        # 6. 导出最终数据
        print(f"\n📊 爬取完成！共爬取 {current_page_num} 页，累计获取 {len(all_data)} 条目标日期({target_date})有效数据")
        if all_data:
            export_to_json(all_data)  # 新增JSON导出
        else:
            print("❌ 无有效目标数据，不生成导出文件")
        
    except Exception as e:
        print(f"❌ 程序异常: {e}")
        # 异常时自动备份已爬取的数据
        if 'all_data' in locals() and all_data:
            backup_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_json = f"重庆市公共资源交易数据_异常备份_{backup_timestamp}.json"
            export_to_json(all_data, backup_json)
    finally:
        if driver:
            print("\n🔌 关闭浏览器")
            driver.quit()

if __name__ == "__main__":
    main()



