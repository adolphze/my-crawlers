import json
import time
import os
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

# ===================== 配置参数 =====================
TARGET_URL = "http://ztb.guizhou.gov.cn/trade/?category=affiche"
BASE_URL = "https://ztb.guizhou.gov.cn"
# 保持跨平台统一字段规范，完全兼容历史版本
FIELDS = [
    "标题", "链接", "发布日期", "省份", "来源平台",
    "业务类型", "信息类型", "行业"
]
# 页面等待与重试配置
PAGE_WAIT = 2
MAX_RETRY = 3
ELEMENT_WAIT_TIMEOUT = 15  # 元素加载超时时间，适配网络波动

# 项目类型-业务类型映射（从页面HTML提取，自动补全空字段）
BUSINESS_TYPE_MAP = {
    "A": "工程建设",
    "B": "土地使用和矿业权",
    "C": "国有产权",
    "D": "政府采购",
    "D4": "药品采购",
    "Z": "货物与服务"
}

# ===================== 无头浏览器初始化（反爬增强）=====================
def init_driver():
    """初始化无头Edge浏览器，适配GitHub Actions Ubuntu环境，绕过反爬检测"""
    edge_options = Options()
    
    # 无头模式核心配置
    edge_options.add_argument("--headless=new")  # Edge新版无头模式，降低被检测概率
    edge_options.add_argument("--no-sandbox")  # Ubuntu环境必须参数
    edge_options.add_argument("--disable-dev-shm-usage")  # 解决共享内存不足问题
    edge_options.add_argument("--disable-gpu")  # 无头环境禁用GPU
    edge_options.add_argument("--window-size=1920,1080")  # 固定窗口尺寸，规避分辨率检测
    
    # 反爬核心配置：隐藏自动化特征
    edge_options.add_argument("--disable-blink-features=AutomationControlled")
    edge_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    edge_options.add_experimental_option("useAutomationExtension", False)
    
    # 真实浏览器指纹配置
    edge_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0")
    edge_options.add_argument("--accept-language=zh-CN,zh;q=0.9,en;q=0.8")
    edge_options.add_argument("--disable-extensions")
    edge_options.add_argument("--disable-plugins-discovery")
    
    # 启动浏览器
    driver = webdriver.Edge(options=edge_options)
    driver.maximize_window()
    
    # 执行JS脚本，彻底隐藏selenium自动化特征
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            // 隐藏webdriver核心属性
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            // 伪造浏览器插件信息
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            // 伪造语言环境
            Object.defineProperty(navigator, 'languages', {
                get: () => ['zh-CN', 'zh', 'en']
            });
            // 伪造chrome对象
            window.chrome = { runtime: {} };
            // 伪造权限查询结果
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({
                    query: () => Promise.resolve({ state: 'granted' })
                })
            });
            // 覆盖selenium特征标识
            Object.defineProperty(window, 'cdc_adoQpoasnfa76pfcZLmcfl_', {
                get: () => undefined
            });
        """
    })
    
    print("✅ 无头Edge浏览器已启动，反爬配置已加载")
    return driver

# ===================== 数据解析函数 =====================
def parse_notice_item(notice_soup):
    """解析单个公告条目，保持原有字段完全兼容，自动补全业务/信息类型"""
    item = {field: "" for field in FIELDS}
    item["省份"] = "贵州省"
    
    try:
        # 1. 标题、链接、来源平台提取
        title_td = notice_soup.find("td")
        if title_td:
            a_tag = title_td.find("a")
            if a_tag:
                # 提取完整公告标题
                item["标题"] = a_tag.get("title", "").strip() or a_tag.get_text(strip=True)
                # 补全可直接访问的完整链接
                href = a_tag.get("href", "").strip()
                if href.startswith("/"):
                    item["链接"] = BASE_URL + href
                else:
                    item["链接"] = href
                
                # 从链接路径提取业务类型
                for code, name in BUSINESS_TYPE_MAP.items():
                    if f"prjtype={code}" in href:
                        item["业务类型"] = name
                        break
            
            # 提取来源平台信息
            source_spans = title_td.find_all("span", class_="source")
            for span in source_spans:
                span_text = span.get_text(strip=True)
                if "来源平台：" in span_text:
                    source_value = span.find("span")
                    if source_value:
                        item["来源平台"] = source_value.get_text(strip=True)
                    break
        
        # 2. 信息类型提取（表格第二列）
        td_list = notice_soup.find_all("td")
        if len(td_list) >= 2:
            item["信息类型"] = td_list[1].get_text(strip=True)
        
        # 3. 发布日期提取（表格第四列，处理时分秒，仅保留日期）
        if len(td_list) >= 4:
            full_date = td_list[3].get_text(strip=True)
            item["发布日期"] = full_date.split(" ")[0]  # 拆分时分秒，仅保留YYYY-MM-DD
        
        # 行业字段保留兼容，可按需扩展
        item["行业"] = ""
                
    except Exception as e:
        print(f"⚠️ 解析单条数据失败: {e}")
    
    return item

def parse_current_page(driver):
    """解析当前页面所有公告数据，带重试机制"""
    soup = BeautifulSoup(driver.page_source, "html.parser")
    # 适配贵州省平台的表格结构
    notice_list = soup.select("table.table-hover > tbody > tr")
    page_data = []
    
    for notice in notice_list:
        item = parse_notice_item(notice)
        page_data.append(item)
    
    print(f"📄 当前页解析出 {len(page_data)} 条数据")
    return page_data

def get_total_pages(driver):
    """获取总页数（从页面直接读取，彻底解决漏爬问题）"""
    try:
        total_page_elem = WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, "//ul[@class='pagination']//em[@data-bind='text:totalPage']"))
        )
        total_pages = int(total_page_elem.text.strip())
        print(f"📊 检测到总页数为: {total_pages}")
        return total_pages
    except Exception as e:
        print(f"⚠️ 获取总页数失败: {e}")
    return None

def has_next_page(driver):
    """判断是否存在下一页（适配贵州省平台分页组件）"""
    try:
        next_btn_list = driver.find_elements(By.XPATH, "//ul[@class='pagination']//a[contains(text(),'下一页')]")
        if next_btn_list:
            next_btn = next_btn_list[0]
            # 检查按钮是否可点击（排除禁用状态）
            btn_class = next_btn.get_attribute("class") or ""
            if next_btn.is_enabled() and "dis" not in btn_class and "disabled" not in btn_class:
                return next_btn
        return None
    except Exception as e:
        print(f"⚠️ 检测下一页失败: {e}")
        return None

def go_to_page_by_input(driver, page_num):
    """通过页码输入框直接跳转（兜底翻页方案，防止漏页）"""
    try:
        input_box = driver.find_element(By.XPATH, "//ul[@class='pagination']//input[@class='number']")
        go_btn = driver.find_element(By.XPATH, "//ul[@class='pagination']//a[contains(text(),'转到')]")
        
        input_box.clear()
        input_box.send_keys(str(page_num))
        go_btn.click()
        time.sleep(PAGE_WAIT)
        # 等待页面列表刷新
        WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-hover > tbody > tr"))
        )
        return True
    except Exception as e:
        print(f"⚠️ 跳转至第 {page_num} 页失败: {e}")
        return False

# ===================== Excel导出函数 =====================
def export_to_excel(data, filename=None):
    """导出数据到Excel，统一保存到data/目录，适配GitHub Actions"""
    if not data:
        print("❌ 无数据可导出")
        return
    
    # 生成默认文件名
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"贵州省公告公示数据_{timestamp}.xlsx"
    
    # 确保data目录存在，GitHub Actions自动创建
    os.makedirs("data", exist_ok=True)
    filepath = os.path.join("data", filename)
    
    # 创建工作簿与工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "公告公示数据"
    
    # 设置表头样式（与历史版本完全统一）
    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 写入表头
    for col, field in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    
    # 写入数据行
    for row, item in enumerate(data, 2):
        for col, field in enumerate(FIELDS, 1):
            ws.cell(row=row, column=col, value=item[field])
    
    # 自动调整列宽
    for col in range(1, len(FIELDS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # 保存文件
    wb.save(filepath)
    print(f"✅ 数据已导出到: {filepath}")
    return filepath

# ===================== 主程序（全自动运行，无人工干预）=====================
def main():
    print("="*60)
    print("📢 贵州省公共资源交易平台全自动爬虫启动")
    print("="*60)
    
    # 计算北京时间前一天日期（适配GitHub Actions UTC时区）
    beijing_now = datetime.utcnow() + timedelta(hours=8)
    target_date = (beijing_now - timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"📅 目标爬取日期（前一天）: {target_date}")
    
    all_data = []
    current_page_num = 1
    consecutive_failures = 0
    driver = None
    
    try:
        # 1. 初始化无头浏览器
        driver = init_driver()
        
        # 2. 导航到目标公告页面
        driver.get(TARGET_URL)
        print(f"🌐 已导航到目标公告网址: {TARGET_URL}")
        
        # 3. 等待页面核心元素加载完成
        WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-hover > tbody"))
        )
        print("✅ 页面核心元素加载完成")
        
        # 4. 自动点击「近三天」按钮，缩小数据范围，提升爬取效率
        try:
            three_days_btn = WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.element_to_be_clickable((By.XPATH, "//ul[@data-bind='foreach:pubTimes']//li[span[text()='近三天']]"))
            )
            # 点击近三天筛选按钮
            three_days_btn.click()
            print("✅ 已点击「近三天」筛选按钮")
            
            # 等待页面列表刷新完成
            time.sleep(PAGE_WAIT)
            WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-hover > tbody > tr"))
            )
            print("✅ 数据列表已刷新")
            
        except Exception as e:
            print(f"⚠️ 点击筛选按钮失败，将使用全量数据过滤: {e}")
        
        # 5. 获取总页数
        total_pages = get_total_pages(driver)
        
        # 6. 循环爬取所有页面
        while True:
            print(f"\n🚀 正在爬取第 {current_page_num} 页...")
            
            # 解析当前页，带重试机制
            page_raw_data = []
            for retry in range(MAX_RETRY):
                try:
                    page_raw_data = parse_current_page(driver)
                    if page_raw_data:
                        consecutive_failures = 0
                        break
                    else:
                        print(f"⚠️ 第 {current_page_num} 页数据为空，重试 ({retry+1}/{MAX_RETRY})...")
                        time.sleep(PAGE_WAIT)
                except Exception as e:
                    print(f"⚠️ 解析第 {current_page_num} 页失败，重试 ({retry+1}/{MAX_RETRY}): {e}")
                    time.sleep(PAGE_WAIT)
            
            # 过滤出目标日期（前一天）的数据
            if page_raw_data:
                target_data = [item for item in page_raw_data if item["发布日期"] == target_date]
                all_data.extend(target_data)
                print(f"🔍 本页共 {len(page_raw_data)} 条，目标日期({target_date})数据 {len(target_data)} 条")
            else:
                consecutive_failures += 1
                print(f"❌ 第 {current_page_num} 页多次尝试后仍无数据")
            
            # 终止条件判断
            if consecutive_failures >= MAX_RETRY:
                print("🔚 连续多次无数据，终止爬取")
                break
            
            # 已知总页数的终止判断
            if total_pages and current_page_num >= total_pages:
                print("🔚 已到达总页数，爬取完成")
                break
            
            # 翻页逻辑：优先点击下一页，失败则用输入框兜底跳转
            current_page_num += 1
            flip_success = False
            
            # 方式1：点击下一页
            next_btn = has_next_page(driver)
            if next_btn:
                try:
                    next_btn.click()
                    time.sleep(PAGE_WAIT)
                    # 等待页面列表加载完成
                    WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                        lambda d: len(d.find_elements(By.CSS_SELECTOR, "table.table-hover > tbody > tr")) > 0
                    )
                    flip_success = True
                except Exception as e:
                    print(f"⚠️ 点击下一页失败: {e}")
            
            # 方式2：点击失败，通过输入框直接跳转
            if not flip_success and total_pages:
                print(f"🔄 尝试通过输入框直接跳转至第 {current_page_num} 页...")
                flip_success = go_to_page_by_input(driver, current_page_num)
            
            # 两种方式都失败，终止爬取
            if not flip_success:
                print(f"❌ 无法翻至第 {current_page_num} 页，终止爬取")
                break
        
        # 7. 导出最终数据
        print(f"\n📊 爬取完成！共爬取 {current_page_num-1} 页，累计获取 {len(all_data)} 条目标日期({target_date})数据")
        export_to_excel(all_data)
        
    except Exception as e:
        print(f"❌ 程序异常: {e}")
        # 异常时自动备份已爬取的数据，防止数据丢失
        if 'all_data' in locals() and all_data:
            backup_filename = f"贵州省公告公示数据_异常备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            export_to_excel(all_data, backup_filename)
    finally:
        if driver:
            print("\n🔌 关闭浏览器")
            driver.quit()

if __name__ == "__main__":
    main()



