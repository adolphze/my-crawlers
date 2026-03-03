import json
import time
import re
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

# ===================== 配置参数 =====================
TARGET_URL = "https://ggzyjy.sc.gov.cn/jyxx/transactionInfo.html"
BASE_URL = "https://ggzyjy.sc.gov.cn"
FIELDS = [
    "标题", "链接", "发布日期", "省份", "来源平台",
    "业务类型", "信息类型", "行业"
]
PAGE_WAIT = 2
MAX_RETRY = 3
ELEMENT_WAIT_TIMEOUT = 10

# ===================== 无头浏览器初始化 =====================
def init_driver():
    """初始化无头Edge浏览器（适用于GitHub Actions）"""
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0')
    driver = webdriver.Edge(options=options)
    print("✅ 无头Edge浏览器已启动")
    return driver

# ===================== 解析函数（保持不变）=====================
def parse_notice_item(notice_soup):
    """解析单个公告条目"""
    item = {field: "" for field in FIELDS}
    item["省份"] = "四川省"
    
    try:
        title_p = notice_soup.find("p", class_="clearfix")
        if title_p:
            a_tag = title_p.find("a", class_="l")
            if a_tag:
                item["标题"] = a_tag.get_text(strip=True)
                href = a_tag.get("href", "").strip()
                item["链接"] = BASE_URL + href if href.startswith("/") else href
            
            date_tag = title_p.find("span", class_="fuInfoDate")
            if date_tag:
                item["发布日期"] = date_tag.get_text(strip=True)
        
        info_spans = notice_soup.find_all("span")
        for span in info_spans:
            span_text = span.get_text(strip=True)
            if "来源：" in span_text:
                source_tag = span.find("i", class_="fuZhuanzai")
                if source_tag:
                    item["来源平台"] = source_tag.get_text(strip=True)
            elif "业务类型：" in span_text:
                business_tag = span.find("i", class_="ywlx")
                if business_tag:
                    item["业务类型"] = business_tag.get_text(strip=True)
            elif "信息类型：" in span_text:
                info_type_tag = span.find("i")
                if info_type_tag:
                    item["信息类型"] = info_type_tag.get_text(strip=True)
        
        item["行业"] = ""
                
    except Exception as e:
        print(f"⚠️ 解析单条数据失败: {e}")
    
    return item

def parse_current_page(driver):
    """解析当前页面所有公告"""
    soup = BeautifulSoup(driver.page_source, "html.parser")
    notice_list = soup.select("ul#transactionInfo > li")
    page_data = []
    
    for notice in notice_list:
        item = parse_notice_item(notice)
        page_data.append(item)
    
    print(f"📄 当前页解析出 {len(page_data)} 条原始数据")
    return page_data

def get_total_pages(driver):
    """获取总页数"""
    try:
        total_text = driver.find_element(By.CSS_SELECTOR, "div.hedy_all").text
        match = re.search(r"总计(\d+)页", total_text)
        if match:
            total_pages = int(match.group(1))
            print(f"📊 检测到总页数为: {total_pages}")
            return total_pages
    except Exception as e:
        print(f"⚠️ 获取总页数失败: {e}")
    return None

def go_to_page_by_input(driver, page_num):
    """通过输入框直接跳转到指定页"""
    try:
        input_box = driver.find_element(By.CSS_SELECTOR, "div.m-pagination-jump input[data-page-btn='jump']")
        go_btn = driver.find_element(By.CSS_SELECTOR, "div.m-pagination-jump button[data-page-btn='jump']")
        input_box.clear()
        input_box.send_keys(str(page_num))
        go_btn.click()
        time.sleep(PAGE_WAIT)
        return True
    except Exception as e:
        print(f"⚠️ 跳转至第 {page_num} 页失败: {e}")
        return False

def has_next_page(driver):
    """判断是否有下一页"""
    try:
        next_btn_list = driver.find_elements(By.XPATH, "//div[@id='page']//ul[@class='m-pagination-page']//a[contains(text(),'下一页')]")
        if next_btn_list:
            next_btn = next_btn_list[0]
            if next_btn.is_enabled() and "disabled" not in (next_btn.get_attribute("class") or ""):
                return next_btn
        return None
    except Exception as e:
        print(f"⚠️ 检测下一页失败: {e}")
        return None

def export_to_excel(data, filename=None):
    """导出数据到Excel"""
    if not data:
        print("❌ 无数据可导出")
        return
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"四川省公共资源交易数据_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "爬取数据"
    
    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, field in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    
    for row, item in enumerate(data, 2):
        for col, field in enumerate(FIELDS, 1):
            ws.cell(row=row, column=col, value=item[field])
    
    for col in range(1, len(FIELDS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # 确保输出目录存在（GitHub Actions中需要）
    import os
    os.makedirs("data", exist_ok=True)
    filepath = os.path.join("data", filename)
    wb.save(filepath)
    print(f"✅ 数据已导出到: {filepath}")
    return filepath

# ===================== 主程序（使用“近一天”筛选 + 日期过滤）=====================
def main():
    try:
        driver = init_driver()
        driver.get(TARGET_URL)
        print(f"🌐 已导航到目标网址: {TARGET_URL}")
        
        # 等待页面初始加载完成
        WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, "transactionInfo"))
        )
        
        # ===== 计算北京时间昨天的日期 =====
        # GitHub Actions 默认时区为 UTC，转换为北京时间（UTC+8）
        beijing_now = datetime.utcnow() + timedelta(hours=8)
        yesterday = (beijing_now - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"📅 目标日期（昨天）: {yesterday}")
        
        # ===== 点击“近一天”按钮 =====
        try:
            pre_oneday_btn = WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "dd[data-id='preOneday']"))
            )
            # 如果未激活则点击
            if "active" not in pre_oneday_btn.get_attribute("class"):
                pre_oneday_btn.click()
                print("✅ 已点击“近一天”按钮")
                time.sleep(1)
            else:
                print("ℹ️ “近一天”已是激活状态")
            
            # 等待列表刷新（等待原有第一条数据消失）
            WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.staleness_of(driver.find_element(By.CSS_SELECTOR, "ul#transactionInfo > li:first-child"))
            )
            # 再等待新列表出现
            WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "ul#transactionInfo > li"))
            )
            print("✅ 数据列表已刷新为近一天数据")
            time.sleep(1)
        except Exception as e:
            print(f"⚠️ 点击“近一天”失败，将使用当前页面继续爬取: {e}")
            # 失败时继续，可能爬取的是默认数据
        
        # 获取总页数
        total_pages = get_total_pages(driver)
        if not total_pages:
            print("⚠️ 无法获取总页数，将使用下一页模式")
        
        all_data = []
        current_page_num = 1
        consecutive_failures = 0
        
        while True:
            print(f"\n🚀 正在爬取第 {current_page_num} 页...")
            
            # 解析当前页（带重试）
            page_raw_data = []
            for retry in range(MAX_RETRY):
                try:
                    page_raw_data = parse_current_page(driver)
                    if page_raw_data:
                        consecutive_failures = 0
                        break
                    else:
                        print(f"⚠️ 第 {current_page_num} 页数据为空，重试 ({retry+1}/{MAX_RETRY})...")
                        time.sleep(PAGE_WAIT)
                except Exception as e:
                    print(f"⚠️ 解析第 {current_page_num} 页失败，重试 ({retry+1}/{MAX_RETRY}): {e}")
                    time.sleep(PAGE_WAIT)
            
            # 过滤出发布日期为昨天的数据
            if page_raw_data:
                yesterday_data = [item for item in page_raw_data if item["发布日期"] == yesterday]
                all_data.extend(yesterday_data)
                print(f"🔍 本页共 {len(page_raw_data)} 条，其中昨天({yesterday})的有 {len(yesterday_data)} 条")
            else:
                consecutive_failures += 1
                print(f"❌ 第 {current_page_num} 页多次尝试后仍无数据")
            
            # 终止条件判断
            if total_pages:
                if current_page_num >= total_pages:
                    print("🔚 已到达总页数")
                    break
                
                current_page_num += 1
                success = False
                
                # 尝试点击下一页
                next_btn = has_next_page(driver)
                if next_btn:
                    try:
                        next_btn.click()
                        time.sleep(PAGE_WAIT)
                        WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                            EC.presence_of_element_located((By.ID, "transactionInfo"))
                        )
                        success = True
                    except Exception as e:
                        print(f"⚠️ 点击下一页失败: {e}")
                
                # 备用：通过输入框跳转
                if not success:
                    print(f"🔄 尝试通过输入框直接跳转至第 {current_page_num} 页...")
                    success = go_to_page_by_input(driver, current_page_num)
                
                if not success:
                    print(f"❌ 无法翻至第 {current_page_num} 页，爬取终止")
                    break
            else:
                # 未知总页数模式
                next_btn = has_next_page(driver)
                if next_btn and consecutive_failures < MAX_RETRY:
                    try:
                        next_btn.click()
                        time.sleep(PAGE_WAIT)
                        current_page_num += 1
                    except Exception as e:
                        print(f"⚠️ 点击下一页失败: {e}")
                        break
                else:
                    print("🔚 已爬取到最后一页或连续失败次数过多")
                    break
        
        print(f"\n📊 爬取完成！共爬取 {current_page_num} 页，累计获取 {len(all_data)} 条昨天（{yesterday}）的数据")
        export_to_excel(all_data)
        
    except Exception as e:
        print(f"❌ 程序异常: {e}")
        if 'all_data' in locals() and all_data:
            export_to_excel(all_data, f"四川省公共资源交易数据_异常备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    finally:
        if 'driver' in locals():
            print("\n🔌 关闭浏览器")
            driver.quit()

if __name__ == "__main__":
    main()



